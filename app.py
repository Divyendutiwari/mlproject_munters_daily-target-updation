"""
Munters Panel Production Planning - Flask Application
======================================================
Main web application serving the dashboard.
"""
import sys, os, json, io
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, render_template, jsonify, request, send_file
import pandas as pd
import numpy as np
from datetime import datetime

import config
from data_loader import load_panel_dataset, load_machine_uptime, load_thermal_timing, load_non_thermal_timing, enrich_dataset
from classifier import classify_panels, assign_production_time, get_class_timing_map
from scheduler import create_schedule, get_schedule_summary, schedule_to_dataframe
from ml_engine import ProductionMLEngine

from database import init_database, save_panels, save_schedule, save_ml_metrics

app = Flask(__name__)

# Global data store
APP_DATA = {}

def format_excel_ws_as_table(ws, df, table_name):
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    
    max_row = len(df) + 1
    max_col = len(df.columns)
    if max_row <= 1 or max_col == 0:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    safe_name = "".join(c for c in table_name if c.isalnum())
    if not safe_name: safe_name = "Table1"
    
    tab = Table(displayName=safe_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def initialize_system(data_path=None):
    """Run the full pipeline: load, classify, ML, schedule, save."""
    print("\n" + "="*60)
    print("  MUNTERS PRODUCTION PLANNING SYSTEM - INITIALIZING")
    print("="*60)
    
    # 1. Initialize database
    try:
        init_database()
    except Exception as e:
        print(f"[DB ERROR] Could not initialize database: {e}")
    
    # 2. Load data
    print("\n--- STEP 1: Loading Data ---")
    try:
        df = load_panel_dataset(data_path)
        machines = load_machine_uptime(data_path)
    except Exception as e:
        print(f"[DATA WARNING] Initial data not found or failed to load: {e}")
        print("[INIT] Starting with an empty state. Please upload an Excel file via the dashboard.")
        APP_DATA["shift_active"] = False
        return

    thermal_timing = load_thermal_timing()
    non_thermal_timing = load_non_thermal_timing()
    
    # 3. Enrich dataset
    print("\n--- STEP 2: Enriching Dataset ---")
    df = enrich_dataset(df)
    
    # 4. Classify panels
    print("\n--- STEP 3: Classifying Panels ---")
    df, area_bounds, length_bounds = classify_panels(df)
    
    # 5. Assign production times
    print("\n--- STEP 4: Assigning Production Times ---")
    df = assign_production_time(df, thermal_timing, non_thermal_timing)
    
    # Initialize live status tracking
    df["Status"] = "Pending"
    
    # 6. Train or Load ML models
    print("\n--- STEP 5: ML Models ---")
    ml_engine = ProductionMLEngine()
    ml_metrics = None
    if os.path.exists(os.path.join(config.MODELS_DIR, "metrics.pkl")):
        ml_metrics = ml_engine.load_models()
        
    if not ml_metrics:
        print("  [ML] Training new models...")
        ml_metrics = ml_engine.train_and_optimize(df, use_fast_grid=True)
    
    # 7. Generate schedule
    print("\n--- STEP 6: Generating Machine Schedule ---")
    schedule_results = create_schedule(df)
    schedule_summary = get_schedule_summary(schedule_results)
    
    # 8. Save to database
    print("\n--- STEP 7: Saving to Database ---")
    save_panels(df)
    save_schedule(schedule_results)
    save_ml_metrics(ml_metrics)
    
    # 9. Store in memory
    APP_DATA["df"] = df
    APP_DATA["machines"] = machines
    APP_DATA["area_bounds"] = area_bounds
    APP_DATA["length_bounds"] = length_bounds
    APP_DATA["schedule"] = schedule_results
    APP_DATA["schedule_summary"] = schedule_summary
    APP_DATA["ml_metrics"] = ml_metrics
    APP_DATA["ml_engine"] = ml_engine
    APP_DATA["thermal_timing"] = thermal_timing
    APP_DATA["non_thermal_timing"] = non_thermal_timing
    APP_DATA["class_timing"] = get_class_timing_map(df)
    APP_DATA["shift_active"] = True
    
    print("\n" + "="*60)
    print("  [OK] SYSTEM INITIALIZED SUCCESSFULLY")
    print(f"  Panels: {len(df)} | Classes: {df['Production_Class'].nunique()}")
    print(f"  Scheduled: {schedule_summary['total_panels_scheduled']} panels")
    print(f"  Tool Changes: {schedule_summary['total_tool_changes']}")
    print(f"  Avg Utilization: {schedule_summary['avg_utilization_pct']}%")
    print("="*60 + "\n")


@app.route("/")
def dashboard():
    return render_template("dashboard.html")


@app.route("/api/shift_status")
def api_shift_status():
    return jsonify({
        "shift_active": APP_DATA.get("shift_active", False)
    })


@app.route("/api/end_shift", methods=["POST"])
def api_end_shift():
    APP_DATA["shift_active"] = False
    
    df = APP_DATA["df"]
    schedule = APP_DATA["schedule"]
    
    # Build panel → machine map
    panel_machine_map = {}
    for machine, data in schedule.items():
        for entry in data["schedule"]:
            if entry["type"] == "production":
                panel_machine_map[entry["panel_id"]] = machine
    
    # All scheduled panel IDs
    scheduled_ids = set(panel_machine_map.keys())
    
    # Backlog: NOT completed panels
    backlog_df = df[df["Status"] != "Completed"].copy()
    backlog_df["Scheduled_Machine"] = backlog_df["Panel_ID"].apply(
        lambda pid: panel_machine_map.get(pid, "Unscheduled"))
    backlog_df["Backlog_Type"] = backlog_df["Panel_ID"].apply(
        lambda pid: "Scheduled_Undone" if pid in scheduled_ids else "Unscheduled")
    
    # Store backlog for next-day merging
    APP_DATA["backlog"] = backlog_df.copy()
    
    # Generate styled Excel
    from openpyxl.styles import PatternFill
    export_cols = ["FG_Design_Code", "Panel_Type", "Production_Class",
                   "Length_mm", "Breadth_mm", "Area_mm2",
                   "Scheduled_Machine", "Backlog_Type"]
    existing = [c for c in export_cols if c in backlog_df.columns]
    export_df = backlog_df[existing].reset_index(drop=True)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Backlog")
        ws = writer.sheets["Backlog"]
        format_excel_ws_as_table(ws, export_df, "BacklogTable")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        bt_col = list(export_df.columns).index("Backlog_Type") + 2  # +2 for 1-index + header
        for row_idx in range(2, len(export_df) + 2):
            bt_val = ws.cell(row=row_idx, column=bt_col - 1 + 1).value
            fill = red_fill if bt_val == "Scheduled_Undone" else yellow_fill
            for col_idx in range(1, len(existing) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        # Write Completed Panels
        completed_df = df[df["Status"] == "Completed"].copy()
        if not completed_df.empty:
            # Done in a day sheet calculation
            completed_df["Strokes"] = completed_df["Panel_Type"].apply(lambda x: 12 if x == "Thermal" else 4)
            
            done_summary = completed_df.groupby(["FG_Design_Code", "Panel_Type"]).agg(
                Total_Panels=("Panel_ID", "count"),
                Total_Area_mm2=("Area_mm2", "sum"),
                Total_Strokes=("Strokes", "sum")
            ).reset_index()
            
            done_summary.to_excel(writer, index=False, sheet_name="Done_in_a_Day")
            ws_done = writer.sheets["Done_in_a_Day"]
            format_excel_ws_as_table(ws_done, done_summary, "DoneInADayTable")
            
            # Add grand totals at the bottom
            max_row = len(done_summary) + 2
            ws_done.cell(row=max_row, column=1, value="GRAND TOTAL")
            ws_done.cell(row=max_row, column=3, value=done_summary["Total_Panels"].sum())
            ws_done.cell(row=max_row, column=4, value=done_summary["Total_Area_mm2"].sum())
            ws_done.cell(row=max_row, column=5, value=done_summary["Total_Strokes"].sum())
            
            green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
            for row_idx in range(2, len(done_summary) + 2):
                for col_idx in range(1, len(done_summary.columns) + 1):
                    ws_done.cell(row=row_idx, column=col_idx).fill = green_fill
                    
            # Completed Panels sheet
            comp_cols = ["FG_Design_Code", "Panel_Type", "Production_Class", "Length_mm", "Breadth_mm", "Area_mm2", "Strokes"]
            comp_existing = [c for c in comp_cols if c in completed_df.columns]
            comp_export = completed_df[comp_existing].reset_index(drop=True)
            comp_export.to_excel(writer, index=False, sheet_name="Completed_Panels")
            
            ws_comp = writer.sheets["Completed_Panels"]
            format_excel_ws_as_table(ws_comp, comp_export, "CompletedTable")
            for row_idx in range(2, len(comp_export) + 2):
                for col_idx in range(1, len(comp_existing) + 1):
                    ws_comp.cell(row=row_idx, column=col_idx).fill = green_fill
                    
    output.seek(0)
    
    date_str = datetime.now().strftime("%Y-%m-%d")
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"End_Of_Shift_{date_str}.xlsx")


@app.route("/api/kpis")
def api_kpis():
    if "df" not in APP_DATA:
        return jsonify({
            "total_orders": 0, "thermal_panels": 0, "non_thermal_panels": 0, "total_classes": 0,
            "total_scheduled": 0, "total_tool_changes": 0, "completed_panels": 0, "pending_panels": 0,
            "total_strokes_done": 0, "total_area_done": 0,
            "best_model": "None", "best_r2": 0, "best_mae": 0, "shift_capacity": config.EFFECTIVE_CAPACITY_MINUTES,
        })

    df = APP_DATA["df"]
    ss = APP_DATA["schedule_summary"]
    ml = APP_DATA["ml_metrics"]
    best_model = APP_DATA["ml_engine"].best_model_name
    
    completed_df = df[df["Status"] == "Completed"]
    strokes_done = 0
    area_done = 0
    if not completed_df.empty:
        strokes_done += len(completed_df[completed_df["Panel_Type"] == "Thermal"]) * 12
        strokes_done += len(completed_df[completed_df["Panel_Type"] == "Non-Thermal"]) * 4
        area_done = int(completed_df["Area_mm2"].sum())
    
    return jsonify({
        "total_orders": len(df),
        "thermal_panels": int((df["Panel_Type"] == "Thermal").sum()),
        "non_thermal_panels": int((df["Panel_Type"] == "Non-Thermal").sum()),
        "total_classes": int(df["Production_Class"].nunique()),
        "total_scheduled": ss["total_panels_scheduled"],
        "total_tool_changes": ss["total_tool_changes"],
        "completed_panels": int(len(completed_df)),
        "pending_panels": int((df["Status"] == "Pending").sum()),
        "total_strokes_done": int(strokes_done),
        "total_area_done": area_done,
        "best_model": best_model,
        "best_r2": round(ml[best_model]["r2"] * 100, 2),
        "best_mae": round(ml[best_model]["mae"], 2),
        "shift_capacity": config.EFFECTIVE_CAPACITY_MINUTES,
    })


@app.route("/api/class_distribution")
def api_class_distribution():
    if "df" not in APP_DATA: return jsonify([])
    df = APP_DATA["df"]
    class_stats = df.groupby("Production_Class").agg(
        count=("FG_Design_Code", "count"),
        avg_area=("Area_mm2", "mean"),
        avg_length=("Length_mm", "mean"),
        avg_time=("Production_Time_Sec", "mean"),
        total_time=("Production_Time_Sec", "sum"),
    ).reset_index()
    class_stats.columns = ["class_name", "panel_count", "avg_area", "avg_length", "avg_time_sec", "total_time_sec"]
    class_stats["avg_area"] = class_stats["avg_area"].round(0).astype(int)
    class_stats["avg_time_sec"] = class_stats["avg_time_sec"].round(1)
    class_stats["total_time_min"] = (class_stats["total_time_sec"] / 60).round(1)
    return jsonify(class_stats.to_dict(orient="records"))


@app.route("/api/schedule")
def api_schedule():
    if "schedule" not in APP_DATA: return jsonify({})
    schedule = APP_DATA["schedule"]
    df = APP_DATA["df"]
    result = {}
    for machine, data in schedule.items():
        annotated_schedule = []
        for entry in data["schedule"]:
            entry_copy = entry.copy()
            if entry["type"] == "production":
                match = df.loc[df["Panel_ID"] == entry["panel_id"], "Status"]
                entry_copy["status"] = match.iloc[0] if len(match) > 0 else "Pending"
            else:
                entry_copy["status"] = "Pending"
            annotated_schedule.append(entry_copy)
            
        result[machine] = {
            "schedule": annotated_schedule,
            "stats": {
                "total_time_used": data["total_time_used"],
                "production_time": data["production_time"],
                "tool_change_time": data["tool_change_time"],
                "idle_time": data["idle_time"],
                "panels_produced": data["panels_produced"],
                "tool_changes": data["tool_changes"],
                "utilization_pct": data["utilization_pct"],
                "capacity_min": data["capacity_min"],
            }
        }
    return jsonify(result)


@app.route("/api/charts/area_distribution")
def api_area_dist():
    if "df" not in APP_DATA: return jsonify({"thermal": [], "non_thermal": []})
    df = APP_DATA["df"]
    thermal = df[df["Panel_Type"] == "Thermal"]["Area_mm2"].tolist()
    non_thermal = df[df["Panel_Type"] == "Non-Thermal"]["Area_mm2"].tolist()
    return jsonify({"thermal": thermal, "non_thermal": non_thermal})


@app.route("/api/charts/panel_type_split")
def api_panel_split():
    if "df" not in APP_DATA: return jsonify({})
    df = APP_DATA["df"]
    counts = df["Panel_Type"].value_counts().to_dict()
    return jsonify(counts)


@app.route("/api/charts/class_counts")
def api_class_counts():
    if "df" not in APP_DATA: return jsonify({})
    df = APP_DATA["df"]
    counts = df["Production_Class"].value_counts().to_dict()
    return jsonify(counts)


@app.route("/api/charts/machine_utilization")
def api_machine_util():
    if "schedule" not in APP_DATA: return jsonify({})
    schedule = APP_DATA["schedule"]
    result = {}
    for machine, data in schedule.items():
        result[machine] = {
            "production": data["production_time"],
            "tool_change": data["tool_change_time"],
            "idle": data["idle_time"],
            "utilization": data["utilization_pct"],
        }
    return jsonify(result)


@app.route("/api/charts/gantt")
def api_gantt():
    if "schedule" not in APP_DATA: return jsonify([])
    schedule = APP_DATA["schedule"]
    df = APP_DATA["df"]
    gantt_data = []
    for machine, data in schedule.items():
        for entry in data["schedule"]:
            status = "Pending"
            if entry["type"] == "production":
                match = df.loc[df["Panel_ID"] == entry["panel_id"], "Status"]
                if len(match) > 0:
                    status = match.iloc[0]
            gantt_data.append({
                "machine": machine,
                "type": entry["type"],
                "start": entry["start_min"],
                "end": entry["end_min"],
                "class": entry["class"],
                "fg_code": entry["fg_code"],
                "panel_id": entry.get("panel_id", ""),
                "duration": entry["duration_min"],
                "status": status,
            })
    return jsonify(gantt_data)


@app.route("/api/gantt_classes")
def api_gantt_classes():
    """Return class-level info with completion status for the Gantt controls."""
    if "schedule" not in APP_DATA: return jsonify([])
    df = APP_DATA["df"]
    schedule = APP_DATA["schedule"]
    scheduled_classes = set()
    for machine, data in schedule.items():
        for entry in data["schedule"]:
            if entry["type"] == "production":
                scheduled_classes.add(entry["class"])
    class_info = []
    for cls in sorted(scheduled_classes):
        cls_df = df[df["Production_Class"] == cls]
        total = len(cls_df)
        completed = int((cls_df["Status"] == "Completed").sum())
        class_info.append({
            "class_name": cls,
            "total_panels": int(total),
            "completed": completed,
            "pending": int(total - completed),
            "is_completed": completed == total,
            "avg_time_sec": round(cls_df["Production_Time_Sec"].mean(), 1),
            "total_time_min": round(cls_df["Production_Time_Sec"].sum() / 60, 1),
            "panel_type": "Thermal" if cls.startswith("Thermal") else "Non-Thermal",
        })
    return jsonify(class_info)


@app.route("/api/mark_class_completed", methods=["POST"])
def api_mark_class_completed():
    """Mark ALL panels of a given production class as Completed."""
    data = request.json
    class_name = data.get("class_name")
    if not class_name:
        return jsonify({"success": False, "message": "Missing class_name"}), 400
    df = APP_DATA["df"]
    mask = df["Production_Class"] == class_name
    count = int(mask.sum())
    if count == 0:
        return jsonify({"success": False, "message": f"No panels found for class {class_name}"}), 404
    df.loc[mask, "Status"] = "Completed"
    APP_DATA["df"] = df
    return jsonify({"success": True, "count": count, "message": f"{count} panels of class '{class_name}' marked completed"})


@app.route("/api/mark_class_completed_and_sync", methods=["POST"])
def api_mark_class_completed_and_sync():
    """Mark ALL panels of a given production class as Completed and sync to GitHub."""
    data = request.json
    class_name = data.get("class_name")
    if not class_name:
        return jsonify({"success": False, "message": "Missing class_name"}), 400
    
    df = APP_DATA["df"]
    mask = df["Production_Class"] == class_name
    count = int(mask.sum())
    if count == 0:
        return jsonify({"success": False, "message": f"No panels found for class {class_name}"}), 404
        
    df.loc[mask, "Status"] = "Completed"
    APP_DATA["df"] = df
    
    # Save the updated panels to db so status persists
    try:
        from database import save_panels
        save_panels(df)
    except Exception as e:
        print("Database save error:", e)

    # Sync to GitHub
    import subprocess
    git_msg = "Successfully synced to GitHub"
    try:
        # Commit the database changes or any file changes
        cwd = config.BASE_DIR if hasattr(config, 'BASE_DIR') else os.path.dirname(os.path.abspath(__file__))
        subprocess.run(["git", "add", "."], check=True, cwd=cwd)
        # Only commit if there are changes
        status_res = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, cwd=cwd)
        if status_res.stdout.strip():
            subprocess.run(["git", "commit", "-m", f"Dashboard update: Marked class {class_name} as done"], check=True, cwd=cwd)
            subprocess.run(["git", "push"], check=True, cwd=cwd)
        else:
            git_msg = "No changes to sync"
    except Exception as e:
        git_msg = f"Git sync error: {str(e)}"
        print(git_msg)
        
    return jsonify({
        "success": True, 
        "count": count, 
        "message": f"{count} panels of class '{class_name}' marked completed",
        "git_msg": git_msg
    })


@app.route("/api/download_class_excel/<path:class_name>")
def api_download_class_excel(class_name):
    """Download an Excel file containing all panels of a specific class."""
    df = APP_DATA["df"]
    class_df = df[df["Production_Class"] == class_name].copy()
    if class_df.empty:
        return jsonify({"error": f"No panels found for class '{class_name}'"}), 404
    export_cols = ["Panel_ID", "FG_Design_Code", "Panel_Type", "Length_mm", "Breadth_mm",
                   "Area_mm2", "Area_Group", "Length_Group", "Production_Class",
                   "Production_Time_Sec", "Production_Time_Min", "Status"]
    existing = [c for c in export_cols if c in class_df.columns]
    class_df = class_df[existing]
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = class_name[:31]
        class_df.to_excel(writer, index=False, sheet_name=sheet_name)
        format_excel_ws_as_table(writer.sheets[sheet_name], class_df, f"ClassTable{sheet_name}")
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"{class_name}_panels.xlsx")


@app.route("/api/ml_metrics")
def api_ml_metrics():
    if "ml_engine" not in APP_DATA: 
        return jsonify({"comparison": "", "best_model": "None", "metrics": {}})
    ml = APP_DATA["ml_engine"]
    return jsonify({
        "comparison": ml.get_model_comparison(),
        "best_model": ml.best_model_name,
        "metrics": {
            name: {
                "r2": round(m["r2"], 4),
                "mae": round(m["mae"], 2),
                "rmse": round(m["rmse"], 2),
            }
            for name, m in ml.metrics.items()
        }
    })


@app.route("/api/box_simulation")
def api_box_simulation():
    if "area_bounds" not in APP_DATA: return jsonify([])
    df = APP_DATA["df"]
    ab = APP_DATA["area_bounds"]
    
    boxes = []
    for ag in config.AREA_GROUPS:
        lo, hi = ab[ag]
        avg_area = (lo + hi) / 2
        # Scale box size proportionally
        max_area = ab["XXL"][1]
        scale = (avg_area / max_area) ** 0.5  # sqrt for visual proportionality
        
        thermal_count = len(df[(df["Area_Group"] == ag) & (df["Panel_Type"] == "Thermal")])
        non_thermal_count = len(df[(df["Area_Group"] == ag) & (df["Panel_Type"] == "Non-Thermal")])
        
        boxes.append({
            "group": ag,
            "scale": round(scale, 3),
            "area_range": f"{lo:,.0f} - {hi:,.0f}",
            "thermal_count": thermal_count,
            "non_thermal_count": non_thermal_count,
            "total": thermal_count + non_thermal_count,
        })
    
    return jsonify(boxes)


@app.route("/api/mark_completed", methods=["POST"])
def api_mark_completed():
    data = request.json
    panel_id = data.get("panel_id")
    if panel_id:
        df = APP_DATA["df"]
        df.loc[df["Panel_ID"] == panel_id, "Status"] = "Completed"
        APP_DATA["df"] = df
        return jsonify({"success": True, "message": f"Panel {panel_id} marked completed"})
    return jsonify({"success": False, "message": "Missing panel_id"}), 400
@app.route("/api/reschedule", methods=["POST"])
def api_reschedule():
    df = APP_DATA["df"]
    pending_df = df[df["Status"] == "Pending"].copy()
    if pending_df.empty:
        return jsonify({"success": False, "message": "No pending panels to reschedule"}), 400
    
    # Calculate real-time offset from shift start (10:00 AM)
    now = datetime.now()
    shift_start = now.replace(hour=10, minute=0, second=0, microsecond=0)
    elapsed_min = max(0, (now - shift_start).total_seconds() / 60)
    elapsed_min = min(elapsed_min, config.EFFECTIVE_CAPACITY_MINUTES)
    
    print(f"\n[LIVE] Rescheduling {len(pending_df)} pending panels from {elapsed_min:.0f} min offset...")
    schedule_results = create_schedule(pending_df, start_offset_min=elapsed_min)
    schedule_summary = get_schedule_summary(schedule_results)
    APP_DATA["schedule"] = schedule_results
    APP_DATA["schedule_summary"] = schedule_summary
    save_schedule(schedule_results)
    return jsonify({"success": True, "summary": schedule_summary})


@app.route("/api/upload", methods=["POST"])
def api_upload():
    if APP_DATA.get("shift_active", False):
        return jsonify({"success": False, "message": "Cannot upload during active shift. End shift first."}), 400
    
    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file part in the request"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "message": "No file selected"}), 400
        
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        return jsonify({"success": False, "message": "Invalid file format. Please upload an Excel file (.xlsx)"}), 400
    
    from werkzeug.utils import secure_filename
    filename = secure_filename(file.filename)
    filepath = os.path.join(config.UPLOAD_FOLDER, filename)
    
    try:
        file.save(filepath)
        print(f"\n[UPLOAD] New data file uploaded: {filepath}")
        
        # Check if there's a backlog from previous shift to merge
        prev_backlog = APP_DATA.get("backlog", None)
        
        # Initialize system with new data
        initialize_system(data_path=filepath)
        
        next_day_plan_path = None
        
        if prev_backlog is not None and len(prev_backlog) > 0:
            print(f"[UPLOAD] Merging {len(prev_backlog)} backlog panels with new data...")
            
            # Get the raw columns needed from backlog
            backlog_raw_cols = ["FG_Design_Code", "Panel_Type", "Length_mm", "Breadth_mm", "Area_mm2"]
            existing_cols = [c for c in backlog_raw_cols if c in prev_backlog.columns]
            backlog_for_merge = prev_backlog[existing_cols].copy()
            backlog_for_merge["_Source"] = prev_backlog["Backlog_Type"].values  # Scheduled_Undone or Unscheduled
            
            # Get fresh data
            new_df = APP_DATA["df"].copy()
            new_count = len(new_df)
            new_df["_Source"] = "New"
            
            # Combine backlog + new data
            combined_raw = pd.concat([backlog_for_merge, new_df[existing_cols + ["_Source"]]], ignore_index=True)
            
            # Re-run full pipeline on combined data
            from classifier import classify_panels, assign_production_time
            combined_raw.dropna(subset=["Length_mm", "Breadth_mm", "Area_mm2", "Panel_Type"], inplace=True)
            combined_raw.reset_index(drop=True, inplace=True)
            combined_raw["Panel_ID"] = "PID-" + combined_raw.index.astype(str)
            combined_df = enrich_dataset(combined_raw.copy())
            combined_df, area_bounds, length_bounds = classify_panels(combined_df)
            thermal_timing = APP_DATA["thermal_timing"]
            non_thermal_timing = APP_DATA["non_thermal_timing"]
            combined_df = assign_production_time(combined_df, thermal_timing, non_thermal_timing)
            combined_df["Status"] = "Pending"
            
            # Store the source info before overwriting
            sources = combined_raw["_Source"].values
            
            # Schedule the combined dataset
            schedule_results = create_schedule(combined_df)
            schedule_summary = get_schedule_summary(schedule_results)
            
            # Build scheduled panel IDs set
            scheduled_ids = set()
            for machine, data in schedule_results.items():
                for entry in data["schedule"]:
                    if entry["type"] == "production":
                        scheduled_ids.add(entry["panel_id"])
            
            combined_df["Scheduled_Today"] = combined_df["Panel_ID"].apply(
                lambda pid: "✓" if pid in scheduled_ids else "✗")
            combined_df["_Source"] = sources[:len(combined_df)]
            
            # Update APP_DATA
            APP_DATA["df"] = combined_df
            APP_DATA["area_bounds"] = area_bounds
            APP_DATA["length_bounds"] = length_bounds
            APP_DATA["schedule"] = schedule_results
            APP_DATA["schedule_summary"] = schedule_summary
            APP_DATA["class_timing"] = get_class_timing_map(combined_df)
            save_panels(combined_df)
            save_schedule(schedule_results)
            
            # Generate styled Next_Day_Plan.xlsx
            from openpyxl.styles import PatternFill
            plan_cols = ["FG_Design_Code", "Panel_Type", "Production_Class",
                         "Length_mm", "Breadth_mm", "Area_mm2", "Scheduled_Today", "_Source"]
            plan_existing = [c for c in plan_cols if c in combined_df.columns]
            plan_df = combined_df[plan_existing].copy()
            plan_df.rename(columns={"_Source": "Origin"}, inplace=True)
            
            plan_output = io.BytesIO()
            with pd.ExcelWriter(plan_output, engine="openpyxl") as writer:
                plan_df.to_excel(writer, index=False, sheet_name="Next_Day_Plan")
                ws = writer.sheets["Next_Day_Plan"]
                format_excel_ws_as_table(ws, plan_df, "NextDayPlanTable")
                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                origin_col_idx = list(plan_df.columns).index("Origin") + 1
                for row_idx in range(2, len(plan_df) + 2):
                    origin_val = ws.cell(row=row_idx, column=origin_col_idx).value
                    fill = None
                    if origin_val == "Scheduled_Undone":
                        fill = red_fill
                    elif origin_val == "Unscheduled":
                        fill = yellow_fill
                    if fill:
                        for col_idx in range(1, len(plan_df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
            plan_output.seek(0)
            
            date_str = datetime.now().strftime("%Y-%m-%d")
            plan_path = os.path.join(config.UPLOAD_FOLDER, f"Next_Day_Plan_{date_str}.xlsx")
            with open(plan_path, "wb") as f:
                f.write(plan_output.read())
            
            print(f"[UPLOAD] Next Day Plan saved: {plan_path} ({len(combined_df)} total panels)")
            return jsonify({
                "success": True,
                "message": f"Merged {len(prev_backlog)} backlog + {new_count} new = {len(combined_df)} total panels. System re-initialized.",
                "has_plan": True,
            })
        
        return jsonify({"success": True, "message": "File uploaded and system successfully re-initialized.", "has_plan": False})
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[UPLOAD ERROR] {str(e)}")
        return jsonify({"success": False, "message": f"Error processing file: {str(e)}"}), 500


@app.route("/api/download_next_day_plan")
def api_download_next_day_plan():
    date_str = datetime.now().strftime("%Y-%m-%d")
    plan_path = os.path.join(config.UPLOAD_FOLDER, f"Next_Day_Plan_{date_str}.xlsx")
    if os.path.exists(plan_path):
        return send_file(plan_path, as_attachment=True, download_name=f"Next_Day_Plan_{date_str}.xlsx")
    return jsonify({"success": False, "message": "Next day plan not found"}), 404


if __name__ == "__main__":
    initialize_system()
    print(f"\n>> Dashboard running at http://{config.FLASK_HOST}:{config.FLASK_PORT}")
    app.run(host=config.FLASK_HOST, port=config.FLASK_PORT, debug=False)
